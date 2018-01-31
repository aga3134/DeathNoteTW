# -*- coding: utf-8 -*-
"""
Created on Wed Dec 27 17:51:03 2017

@author: aga
"""

from openpyxl import load_workbook
import Util as util
import re

class DataDisease:
    def __init__(self, connection):
        self.connection = connection
        
    def CreateTable(self):
        print("Create Table for DataDisease")
        with self.connection.cursor() as cursor:
            sql = "CREATE TABLE IF NOT EXISTS MedicineStatisticByAge (\
                year INT(11),\
                mType VARCHAR(255),\
                disease VARCHAR(255),\
                subDisease VARCHAR(255),\
                sex VARCHAR(10),\
                minAge FLOAT,\
                maxAge FLOAT,\
                caseNum FLOAT,\
                expense FLOAT,\
                PRIMARY KEY (mType,sex,disease,subDisease,year,minAge,maxAge)\
                );"
            cursor.execute(sql)
            
            """sql = "CREATE TABLE IF NOT EXISTS HealthCareCashFlow (\
                year INT(11),\
                income INT(11),\
                expense INT(11),\
                loan INT(11),\
                balance INT(11),\
                PRIMARY KEY (year)\
                );"
            cursor.execute(sql)
            
            sql = "CREATE TABLE IF NOT EXISTS HealthCareProfit (\
                year INT(11),\
                unitCount INT(11),\
                targetCount INT(11),\
                targetReceivable INT(11),\
                unitReceivable INT(11),\
                govReceivable INT(11),\
                revenue INT(11),\
                cost INT(11),\
                badDebt INT(11),\
                other INT(11),\
                profit INT(11),\
                cumulate INT(11),\
                PRIMARY KEY (year)\
                );"
            cursor.execute(sql)
            
            sql = "CREATE TABLE IF NOT EXISTS HealthCareTarget (\
                year INT(11),\
                targetType INT(11),\
                count INT(11),\
                avgInsurance INT(11),\
                PRIMARY KEY (year,targetType)\
                );"
            cursor.execute(sql)
            
            sql = "CREATE TABLE IF NOT EXISTS HealthCareOrganization (\
                year INT(11),\
                attrGroup VARCHAR(255),\
                attr VARCHAR(255),\
                count INT(11),\
                PRIMARY KEY (year)\
                );"
            cursor.execute(sql)
            
            sql = "CREATE TABLE IF NOT EXISTS HealthCarePayment (\
                year INT(11),\
                attrGroup VARCHAR(255),\
                attr VARCHAR(255),\
                caseNum INT(11),\
                expense INT(11),\
                ratio FLOAT,\
                day INT(11),\
                PRIMARY KEY (year)\
                );"
            cursor.execute(sql)"""

        self.connection.commit()
        
    
        
    def UpdateWestMedicineByAge(self):
        print("西醫門診")
        westCase = load_workbook(filename = 'data/病之章/5.1-20_西醫門診件數--按疾病別性別及年齡.XLSX')
        westCost = load_workbook(filename = 'data/病之章/5.1-21+西醫門診醫療費用--按疾病別性別及年齡.XLSX')
        for year in range(87,105):
            caseSheet = westCase[str(year)]
            expenseSheet = westCost[str(year)]
            d = {"year":year+1911,"mType":"西醫門診"}
            
            for page in range(0,7):
                pageStart = page*14+1
                for offset in range(5,14,3):
                    allCol = pageStart+offset
                    maleCol = pageStart+offset+1
                    femaleCol = pageStart+offset+2
                    minAge = -1
                    maxAge = -1
                    if page == 0:
                        if offset == 5: #總計
                            minAge = 0
                            maxAge = 100
                        elif offset == 8:   #未滿4週
                            minAge = 0
                            maxAge = 0.077
                        elif offset == 11: #4週~1歲
                            minAge = 0.077
                            maxAge = 1
                    elif page == 6: #>85歲
                        if offset == 11:
                            minAge = 85
                            maxAge = 100
                    if minAge == -1 and maxAge == -1:
                        if year > 99:
                            ageRow = 5
                        else:
                            ageRow = 4
                        age = re.findall(r"\d+",caseSheet.cell(row=ageRow, column=allCol).value)
                        minAge = age[0]
                        maxAge = age[1]
                    d["minAge"] = minAge
                    d["maxAge"] = maxAge
                
                    if year > 99:
                        rowStart = 9
                        rowEnd = 53
                    else:
                        rowStart = 8
                        rowEnd = 52
                    for row in range(rowStart,rowEnd):
                        category = caseSheet.cell(row=row, column=1).value
                        disease = caseSheet.cell(row=row, column=2).value
                        subDisease = caseSheet.cell(row=row, column=3).value
                        if category is not None:
                            if "總計" in category:
                                d["disease"] = "總計"
                                d["subDisease"] = "總計"
                            elif "不詳" in category:
                                d["disease"] = "不詳"
                                d["subDisease"] = "不詳"
                            else:
                                disease = disease.split(" ")[0]
                                d["disease"] = disease
                                d["subDisease"] = disease
                        elif subDisease is not None:
                            subDisease = subDisease.split(" ")[0]
                            d["subDisease"] = subDisease
                        d["sex"] = "全部"
                        
                        d["caseNum"] = util.VarifyValue(caseSheet.cell(row=row, column=allCol).value)
                        d["expense"] = util.VarifyValue(expenseSheet.cell(row=row, column=allCol).value)
                        #print(d)
                        util.DataToDB(self.connection,"MedicineStatisticByAge",d)
                        
                        d["sex"] = "男"
                        d["caseNum"] = util.VarifyValue(caseSheet.cell(row=row, column=maleCol).value)
                        d["expense"] = util.VarifyValue(expenseSheet.cell(row=row, column=maleCol).value)
                        #print(d)
                        util.DataToDB(self.connection,"MedicineStatisticByAge",d)
                        
                        d["sex"] = "女"
                        d["caseNum"] = util.VarifyValue(caseSheet.cell(row=row, column=femaleCol).value)
                        d["expense"] = util.VarifyValue(expenseSheet.cell(row=row, column=femaleCol).value)
                        #print(d)
                        util.DataToDB(self.connection,"MedicineStatisticByAge",d)
            self.connection.commit()
            
    def UpdateDentistByAge(self):
        print("牙醫門診")
        dentist = load_workbook(filename = 'data/病之章/5.1-24+健保門診件數與申報費用—牙醫.XLSX')
        for year in range(87,105):
            sheet = dentist[str(year)]
            d = {"year":year+1911,"mType":"牙醫門診"}
            if year > 99:
                rowStart = 8
                rowEnd = 38
            else:
                rowStart = 7
                rowEnd = 35
            for row in range(rowStart,rowEnd):
                disease = sheet.cell(row=row, column=1).value
                subDisease = sheet.cell(row=row, column=2).value
                if disease is not None:
                    disease = disease.split(" ")[0]
                    d["disease"] = disease
                    d["subDisease"] = disease
                elif subDisease is not None:
                    subDisease = subDisease.split(" ")[0]
                    d["subDisease"] = subDisease
                    
                d["minAge"] = 0
                d["maxAge"] = 100
                d["sex"] = "全部"
                d["caseNum"] = util.VarifyValue(sheet.cell(row=row, column=5).value)
                d["expense"] = util.VarifyValue(sheet.cell(row=row, column=18).value)
                #print(d)
                util.DataToDB(self.connection,"MedicineStatisticByAge",d)
                
                d["sex"] = "男"
                d["caseNum"] = util.VarifyValue(sheet.cell(row=row, column=6).value)
                d["expense"] = util.VarifyValue(sheet.cell(row=row, column=19).value)
                #print(d)
                util.DataToDB(self.connection,"MedicineStatisticByAge",d)
                
                d["sex"] = "女"
                d["caseNum"] = util.VarifyValue(sheet.cell(row=row, column=7).value)
                d["expense"] = util.VarifyValue(sheet.cell(row=row, column=20).value)
                #print(d)
                util.DataToDB(self.connection,"MedicineStatisticByAge",d)
                
                d["sex"] = "全部"
                d["minAge"] = 65
                d["maxAge"] = 100
                d["caseNum"] = util.VarifyValue(sheet.cell(row=row, column=13).value)
                d["expense"] = util.VarifyValue(sheet.cell(row=row, column=26).value)
                #print(d)
                util.DataToDB(self.connection,"MedicineStatisticByAge",d)
                
                for col in range(8,13):
                    age = re.findall(r"\d+",sheet.cell(row=6, column=col).value)
                    d["minAge"] = age[0]
                    d["maxAge"] = age[1]
                    d["caseNum"] = util.VarifyValue(sheet.cell(row=row, column=col).value)
                    d["expense"] = util.VarifyValue(sheet.cell(row=row, column=col+13).value)
                    #print(d)
                    util.DataToDB(self.connection,"MedicineStatisticByAge",d)
            self.connection.commit()
    
    def UpdateChineseMedicineByAge(self):
        print("中醫門診")
        cm = load_workbook(filename = 'data/病之章/5.1-25+健保門診件數與申報費用—中醫.XLSX')
        for year in range(87,105):
            sheet = cm[str(year)]
            d = {"year":year+1911,"mType":"中醫門診"}
            if year > 99:
                rowStart = 8
                rowEnd = 52
            else:
                rowStart = 7
                rowEnd = 51
            for row in range(rowStart,rowEnd):
                category = sheet.cell(row=row, column=1).value
                disease = sheet.cell(row=row, column=2).value
                subDisease = sheet.cell(row=row, column=3).value
                if category is not None:
                    if "總計" in category:
                        d["disease"] = "總計"
                        d["subDisease"] = "總計"
                    elif "不詳" in category:
                        d["disease"] = "不詳"
                        d["subDisease"] = "不詳"
                    else:
                        disease = disease.split(" ")[0]
                        d["disease"] = disease
                        d["subDisease"] = disease
                elif subDisease is not None:
                    subDisease = subDisease.split(" ")[0]
                    d["subDisease"] = subDisease
                    
                d["minAge"] = 0
                d["maxAge"] = 100
                d["sex"] = "全部"
                d["caseNum"] = util.VarifyValue(sheet.cell(row=row, column=6).value)
                d["expense"] = util.VarifyValue(sheet.cell(row=row, column=20).value)
                #print(d)
                util.DataToDB(self.connection,"MedicineStatisticByAge",d)
                
                d["sex"] = "男"
                d["caseNum"] = util.VarifyValue(sheet.cell(row=row, column=7).value)
                d["expense"] = util.VarifyValue(sheet.cell(row=row, column=21).value)
                #print(d)
                util.DataToDB(self.connection,"MedicineStatisticByAge",d)
                
                d["sex"] = "女"
                d["caseNum"] = util.VarifyValue(sheet.cell(row=row, column=8).value)
                d["expense"] = util.VarifyValue(sheet.cell(row=row, column=22).value)
                #print(d)
                util.DataToDB(self.connection,"MedicineStatisticByAge",d)
                
                d["sex"] = "全部"
                d["minAge"] = 65
                d["maxAge"] = 100
                d["caseNum"] = util.VarifyValue(sheet.cell(row=row, column=14).value)
                d["expense"] = util.VarifyValue(sheet.cell(row=row, column=28).value)
                #print(d)
                util.DataToDB(self.connection,"MedicineStatisticByAge",d)
                
                for col in range(9,14):
                    age = re.findall(r"\d+",sheet.cell(row=6, column=col).value)
                    d["minAge"] = age[0]
                    d["maxAge"] = age[1]
                    d["caseNum"] = util.VarifyValue(sheet.cell(row=row, column=col).value)
                    d["expense"] = util.VarifyValue(sheet.cell(row=row, column=col+14).value)
                    #print(d)
                    util.DataToDB(self.connection,"MedicineStatisticByAge",d)
            self.connection.commit()
        
    def UpdateEmergencyByAge(self):
        print("急診")
        emCase = load_workbook(filename = 'data/病之章/5.1-26+急診件數--按疾病別性別及年齡.XLSX')
        emCost = load_workbook(filename = 'data/病之章/5.1-27+急診醫療費用--按疾病別性別及年齡.XLSX')
        for year in range(87,105):
            caseSheet = emCase[str(year)]
            expenseSheet = emCost[str(year)]
            d = {"year":year+1911,"mType":"急診"}
            
            for page in range(0,7):
                pageStart = page*14+1
                for offset in range(5,14,3):
                    allCol = pageStart+offset
                    maleCol = pageStart+offset+1
                    femaleCol = pageStart+offset+2
                    minAge = -1
                    maxAge = -1
                    if page == 0:
                        if offset == 5: #總計
                            minAge = 0
                            maxAge = 100
                        elif offset == 8:   #未滿4週
                            minAge = 0
                            maxAge = 0.077
                        elif offset == 11:  #4週~1歲
                            minAge = 0.077
                            maxAge = 1
                    elif page == 6:  #>85歲
                        if offset == 11:
                            minAge = 85
                            maxAge = 100
                    if minAge == -1 and maxAge == -1:
                        if year > 99:
                            ageRow = 5
                        else:
                            ageRow = 4
                        age = re.findall(r"\d+",caseSheet.cell(row=ageRow, column=allCol).value)
                        minAge = age[0]
                        maxAge = age[1]
                    d["minAge"] = minAge
                    d["maxAge"] = maxAge
                
                    if year > 99:
                        rowStart = 9
                        rowEnd = 53
                    else:
                        rowStart = 8
                        rowEnd = 52
                    for row in range(rowStart,rowEnd):
                        category = caseSheet.cell(row=row, column=1).value
                        disease = caseSheet.cell(row=row, column=2).value
                        subDisease = caseSheet.cell(row=row, column=3).value
                        if category is not None:
                            if "總計" in category:
                                d["disease"] = "總計"
                                d["subDisease"] = "總計"
                            elif "不詳" in category:
                                d["disease"] = "不詳"
                                d["subDisease"] = "不詳"
                            else:
                                disease = disease.split(" ")[0]
                                d["disease"] = disease
                                d["subDisease"] = disease
                        elif subDisease is not None:
                            subDisease = subDisease.split(" ")[0]
                            d["subDisease"] = subDisease
                        d["sex"] = "全部"
                        
                        d["caseNum"] = util.VarifyValue(caseSheet.cell(row=row, column=allCol).value)
                        d["expense"] = util.VarifyValue(expenseSheet.cell(row=row, column=allCol).value)
                        #print(d)
                        util.DataToDB(self.connection,"MedicineStatisticByAge",d)
                        
                        d["sex"] = "男"
                        d["caseNum"] = util.VarifyValue(caseSheet.cell(row=row, column=maleCol).value)
                        d["expense"] = util.VarifyValue(expenseSheet.cell(row=row, column=maleCol).value)
                        #print(d)
                        util.DataToDB(self.connection,"MedicineStatisticByAge",d)
                        
                        d["sex"] = "女"
                        d["caseNum"] = util.VarifyValue(caseSheet.cell(row=row, column=femaleCol).value)
                        d["expense"] = util.VarifyValue(expenseSheet.cell(row=row, column=femaleCol).value)
                        #print(d)
                        util.DataToDB(self.connection,"MedicineStatisticByAge",d)
            self.connection.commit()
        
    def UpdateHospitalizedByAge(self):
        print("住院")
        hoCase = load_workbook(filename = 'data/病之章/5.1-28+住院件數—按疾病別性別及年齡.XLSX')
        hoCost = load_workbook(filename = 'data/病之章/5.1-29+住院醫療費用—按疾病別性別及年齡.XLSX')        
        for year in range(87,105):
            caseSheet = hoCase[str(year)]
            expenseSheet = hoCost[str(year)]
            d = {"year":year+1911,"mType":"住院"}
            
            for page in range(0,6):
                pageStart = page*14+1
                if page < 5:
                    offsetEnd = 14
                else:
                    offsetEnd = 17
                for offset in range(5,offsetEnd,3):
                    allCol = pageStart+offset
                    maleCol = pageStart+offset+1
                    femaleCol = pageStart+offset+2
                    minAge = -1
                    maxAge = -1
                    if page == 0:
                        if offset == 5: #總計
                            minAge = 0
                            maxAge = 100
                    elif page == 5: #>85歲
                        if offset == 14:
                            minAge = 85
                            maxAge = 100
                    if minAge == -1 and maxAge == -1:
                        if year > 99:
                            ageRow = 5
                        else:
                            ageRow = 4
                        age = re.findall(r"\d+",caseSheet.cell(row=ageRow, column=allCol).value)
                        minAge = age[0]
                        maxAge = age[1]
                    d["minAge"] = minAge
                    d["maxAge"] = maxAge
                
                    if year > 99:
                        rowStart = 8
                        rowEnd = 52
                    else:
                        rowStart = 7
                        rowEnd = 51
                    for row in range(rowStart,rowEnd):
                        category = caseSheet.cell(row=row, column=1).value
                        disease = caseSheet.cell(row=row, column=2).value
                        subDisease = caseSheet.cell(row=row, column=3).value
                        if category is not None:
                            if "總計" in category:
                                d["disease"] = "總計"
                                d["subDisease"] = "總計"
                            elif "不詳" in category:
                                d["disease"] = "不詳"
                                d["subDisease"] = "不詳"
                            else:
                                disease = disease.split(" ")[0]
                                d["disease"] = disease
                                d["subDisease"] = disease
                        elif subDisease is not None:
                            subDisease = subDisease.split(" ")[0]
                            d["subDisease"] = subDisease
                        d["sex"] = "全部"
                        
                        d["caseNum"] = util.VarifyValue(caseSheet.cell(row=row, column=allCol).value)
                        d["expense"] = util.VarifyValue(expenseSheet.cell(row=row+1, column=allCol).value)
                        #print(d)
                        util.DataToDB(self.connection,"MedicineStatisticByAge",d)
                        
                        d["sex"] = "男"
                        d["caseNum"] = util.VarifyValue(caseSheet.cell(row=row, column=maleCol).value)
                        d["expense"] = util.VarifyValue(expenseSheet.cell(row=row+1, column=maleCol).value)
                        #print(d)
                        util.DataToDB(self.connection,"MedicineStatisticByAge",d)
                        
                        d["sex"] = "女"
                        d["caseNum"] = util.VarifyValue(caseSheet.cell(row=row, column=femaleCol).value)
                        d["expense"] = util.VarifyValue(expenseSheet.cell(row=row+1, column=femaleCol).value)
                        #print(d)
                        util.DataToDB(self.connection,"MedicineStatisticByAge",d)
            self.connection.commit()
            
    def UpdateData(self):
        print("update MedicineStatisticByAge")
        self.UpdateWestMedicineByAge()
        self.UpdateDentistByAge()
        self.UpdateChineseMedicineByAge()
        self.UpdateEmergencyByAge()
        self.UpdateHospitalizedByAge()

