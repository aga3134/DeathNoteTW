# -*- coding: utf-8 -*-
"""
Created on Wed Dec 27 17:51:03 2017

@author: aga
"""

from openpyxl import load_workbook
import Util as util
import re

class DataAging:
    def __init__(self, connection):
        self.connection = connection
        self.wb = load_workbook(filename = 'data/老之章/102老人狀況調查.xlsx')
        
    def CreateTable(self):
        print("Create Table for DataAging")
        with self.connection.cursor() as cursor:
            sql = "CREATE TABLE IF NOT EXISTS AgingSurveyAge (\
                attrGroup VARCHAR(255),\
                attr VARCHAR(255),\
                minAge INT(11),\
                maxAge INT(11),\
                count INT(11),\
                PRIMARY KEY (attrGroup,attr,minAge,maxAge)\
                );"
            cursor.execute(sql)
            
            sql = "CREATE TABLE IF NOT EXISTS AgingSurveyHouseType (\
                attrGroup VARCHAR(255),\
                attr VARCHAR(255),\
                houseType VARCHAR(255),\
                count INT(11),\
                PRIMARY KEY (attrGroup,attr,houseType)\
                );"
            cursor.execute(sql)
            
            sql = "CREATE TABLE IF NOT EXISTS AgingSurveyFamilyType (\
                attrGroup VARCHAR(255),\
                attr VARCHAR(255),\
                familyType VARCHAR(255),\
                count INT(11),\
                PRIMARY KEY (attrGroup,attr,familyType)\
                );"
            cursor.execute(sql)
            
            sql = "CREATE TABLE IF NOT EXISTS AgingSurveyChronic (\
                attrGroup VARCHAR(255),\
                attr VARCHAR(255),\
                chronic VARCHAR(255),\
                count INT(11),\
                PRIMARY KEY (attrGroup,attr,chronic)\
                );"
            cursor.execute(sql)
            
            sql = "CREATE TABLE IF NOT EXISTS AgingSurveyActivity (\
                attrGroup VARCHAR(255),\
                attr VARCHAR(255),\
                activity VARCHAR(255),\
                mainCount INT(11),\
                secondaryCount INT(11),\
                PRIMARY KEY (attrGroup,attr,activity)\
                );"
            cursor.execute(sql)
            
            sql = "CREATE TABLE IF NOT EXISTS AgingSurveyFeeling (\
                attrGroup VARCHAR(255),\
                attr VARCHAR(255),\
                feel VARCHAR(255),\
                neverCount INT(11),\
                sometimesCount INT(11),\
                oftenCount INT(11),\
                PRIMARY KEY (attrGroup,attr,feel)\
                );"
            cursor.execute(sql)
            
            sql = "CREATE TABLE IF NOT EXISTS AgingSurveyIdealLive (\
                attrGroup VARCHAR(255),\
                attr VARCHAR(255),\
                liveType VARCHAR(255),\
                count INT(11),\
                PRIMARY KEY (attrGroup,attr,liveType)\
                );"
            cursor.execute(sql)
            
            sql = "CREATE TABLE IF NOT EXISTS AgingSurveyExpect (\
                attrGroup VARCHAR(255),\
                attr VARCHAR(255),\
                expect VARCHAR(255),\
                count INT(11),\
                PRIMARY KEY (attrGroup,attr,expect)\
                );"
            cursor.execute(sql)
            
            sql = "CREATE TABLE IF NOT EXISTS AgingSurveyLivingDifficulty (\
                attrGroup VARCHAR(255),\
                attr VARCHAR(255),\
                difficulty VARCHAR(255),\
                count INT(11),\
                PRIMARY KEY (attrGroup,attr,difficulty)\
                );"
            cursor.execute(sql)
            
            sql = "CREATE TABLE IF NOT EXISTS AgingSurveyFunctionDifficulty (\
                attrGroup VARCHAR(255),\
                attr VARCHAR(255),\
                difficulty VARCHAR(255),\
                noNeedCount INT(11),\
                canCount INT(11),\
                diffcultCount INT(11),\
                canNotCount INT(11),\
                PRIMARY KEY (attrGroup,attr,difficulty)\
                );"
            cursor.execute(sql)

        self.connection.commit()
        
    def UpdateAgingSurveyAge(self):
        print("update AgingSurveyAge")
        
        sheet = self.wb["受訪者狀況"]
        attrGroup = []
        attrGroup.append({"name":"總計","start":3,'length':1})
        attrGroup.append({"name":"性別","start":5,'length':2})
        attrGroup.append({"name":"身分別","start":8,'length':3})
        attrGroup.append({"name":"教育程度","start":12,'length':7})
        attrGroup.append({"name":"婚姻狀況","start":20,'length':4})
        attrGroup.append({"name":"現有子女數","start":25,'length':5})
        attrGroup.append({"name":"居住方式","start":31,'length':2})
        attrGroup.append({"name":"縣市","start":34,'length':25})
        
        for group in attrGroup:
            d = {"attrGroup":group["name"]}
            for row in range(group["start"],group["start"]+group["length"]):
                #去掉空白
                attr = "".join(sheet['A'+str(row)].value.split())
                total = sheet['B'+str(row)].value
                d["attr"] = attr
                d["minAge"] = 55
                d["maxAge"] = 100
                d["count"] = total
                #print(d)
                util.DataToDB(self.connection,"AgingSurveyAge",d)
                
                for col in range(4,10):
                    age = re.findall(r"\d+",sheet.cell(row=1, column=col).value)
                    d["minAge"] = int(age[0])
                    d["maxAge"] = int(age[1])
                    v = sheet.cell(row=row, column=col).value
                    if util.IsNumber(v):
                        d["count"] = int(total*float(v)/100)
                    else:
                        d["count"] = 0
                    #print(d)
                    util.DataToDB(self.connection,"AgingSurveyAge",d)
                    
                d["minAge"] = 85
                d["maxAge"] = 100
                v = sheet['J'+str(row)].value
                if util.IsNumber(v):  
                    d["count"] = total*float(v)/100
                else:
                    d["count"] = 0
                #print(d)
                util.DataToDB(self.connection,"AgingSurveyAge",d)
    
    def UpdateAgingSurveyHouseType(self):
        print("update AgingSurveyHouseType")
        
        sheet = self.wb["住宅類型"]
        attrGroup = []
        attrGroup.append({"name":"總計","start":3,'length':1})
        attrGroup.append({"name":"性別","start":5,'length':2})
        attrGroup.append({"name":"年齡別","start":8,'length':4})
        attrGroup.append({"name":"身分別","start":13,'length':3})
        attrGroup.append({"name":"教育程度","start":17,'length':7})
        attrGroup.append({"name":"婚姻狀況","start":25,'length':4})
        attrGroup.append({"name":"現有子女數","start":30,'length':5})
        attrGroup.append({"name":"縣市","start":36,'length':25})
        
        for group in attrGroup:
            d = {"attrGroup":group["name"]}
            for row in range(group["start"],group["start"]+group["length"]):
                #去掉空白
                attr = "".join(sheet['A'+str(row)].value.split())
                total = sheet['B'+str(row)].value
                d["attr"] = attr
                d["houseType"] = "總計"
                d["count"] = total
                #print(d)
                util.DataToDB(self.connection,"AgingSurveyHouseType",d)
                
                houseType = []
                houseType.append({"type":"電梯大樓","col":4});
                houseType.append({"type":"公寓-有電梯","col":6});
                houseType.append({"type":"公寓-無電梯","col":7});
                houseType.append({"type":"兩層樓以上家宅(含透天厝、別墅)-有電梯","col":9});
                houseType.append({"type":"兩層樓以上家宅(含透天厝、別墅)-無電梯","col":10});
                houseType.append({"type":"平房(含三合院及四合院)","col":11});
                houseType.append({"type":"一般搭建屋","col":12});
                for house in houseType:
                    col = house["col"]
                    d["houseType"] = house["type"]
                    v = sheet.cell(row=row, column=col).value
                    if util.IsNumber(v):
                        d["count"] = int(total*float(v)/100)
                    else:
                        d["count"] = 0
                    #print(d)
                    util.DataToDB(self.connection,"AgingSurveyHouseType",d)
                    
    
    def UpdateAgingSurveyFamilyType(self):
        print("update AgingSurveyFamilyType")
        
        sheet = self.wb["家庭組成"]
        attrGroup = []
        attrGroup.append({"name":"總計","start":3,'length':1})
        attrGroup.append({"name":"性別","start":5,'length':2})
        attrGroup.append({"name":"年齡別","start":8,'length':4})
        attrGroup.append({"name":"身分別","start":13,'length':3})
        attrGroup.append({"name":"教育程度","start":17,'length':7})
        attrGroup.append({"name":"婚姻狀況","start":25,'length':4})
        attrGroup.append({"name":"現有子女數","start":30,'length':5})
        attrGroup.append({"name":"縣市","start":36,'length':25})
        
        for group in attrGroup:
            d = {"attrGroup":group["name"]}
            for row in range(group["start"],group["start"]+group["length"]):
                #去掉空白
                attr = "".join(sheet['A'+str(row)].value.split())
                total = sheet['B'+str(row)].value
                d["attr"] = attr
                d["familyType"] = "總計"
                d["count"] = total
                #print(d)
                util.DataToDB(self.connection,"AgingSurveyFamilyType",d)
                
                familyType = []
                familyType.append({"type":"獨居","col":4});
                familyType.append({"type":"僅與配偶(同居人)同住","col":5});
                familyType.append({"type":"兩代家庭-與配偶(同居人)及子女同住","col":7});
                familyType.append({"type":"兩代家庭-僅與子女同住","col":8});
                familyType.append({"type":"兩代家庭-與父母同住","col":9});
                familyType.append({"type":"兩代家庭-與(外)孫子女同住","col":10});
                familyType.append({"type":"三代家庭-與子女及(外)孫子女同住","col":12});
                familyType.append({"type":"三代家庭-與父母及子女同住","col":13});
                familyType.append({"type":"三代家庭-與父母及(外)孫子女同住","col":14});
                familyType.append({"type":"四代家庭","col":15});
                familyType.append({"type":"與其他親戚或朋友同住","col":16});
                familyType.append({"type":"僅與外籍看護工同住","col":17});
                familyType.append({"type":"住在機構及其他","col":18});
                for family in familyType:
                    col = family["col"]
                    d["familyType"] = family["type"]
                    v = sheet.cell(row=row, column=col).value
                    if util.IsNumber(v):
                        d["count"] = int(total*float(v)/100)
                    else:
                        d["count"] = 0
                    #print(d)
                    util.DataToDB(self.connection,"AgingSurveyFamilyType",d)
    
    def UpdateAgingSurveyChronic(self):
        print("update AgingSurveyChronic")
        
        sheet = self.wb["慢性病"]
        attrGroup = []
        attrGroup.append({"name":"總計","start":3,'length':1})
        attrGroup.append({"name":"性別","start":5,'length':2})
        attrGroup.append({"name":"年齡別","start":8,'length':4})
        attrGroup.append({"name":"身分別","start":13,'length':3})
        attrGroup.append({"name":"教育程度","start":17,'length':7})
        attrGroup.append({"name":"婚姻狀況","start":25,'length':4})
        attrGroup.append({"name":"現有子女數","start":30,'length':5})
        attrGroup.append({"name":"居住方式","start":36,'length':2})
        attrGroup.append({"name":"縣市","start":39,'length':25})
        
        for group in attrGroup:
            d = {"attrGroup":group["name"]}
            for row in range(group["start"],group["start"]+group["length"]):
                #去掉空白
                attr = "".join(sheet['A'+str(row)].value.split())
                total = sheet['B'+str(row)].value
                d["attr"] = attr
                d["chronic"] = "總計"
                d["count"] = total
                #print(d)
                util.DataToDB(self.connection,"AgingSurveyChronic",d)
                
                for col in range(4,25):
                    #去掉空白
                    chronic = "".join(sheet.cell(row=2, column=col).value.split())
                    d["chronic"] = chronic
                    v = sheet.cell(row=row, column=col).value
                    if util.IsNumber(v):
                        d["count"] = int(total*float(v)/100)
                    else:
                        d["count"] = 0
                    #print(d)
                    util.DataToDB(self.connection,"AgingSurveyChronic",d)
    
    def UpdateAgingSurveyActivity(self):
        print("update AgingSurveyActivity")
        
        sheet = self.wb["日常活動"]
        attrGroup = []
        attrGroup.append({"name":"總計","start":3,'length':1})
        attrGroup.append({"name":"性別","start":5,'length':2})
        attrGroup.append({"name":"年齡別","start":8,'length':4})
        attrGroup.append({"name":"身分別","start":13,'length':3})
        attrGroup.append({"name":"教育程度","start":17,'length':7})
        attrGroup.append({"name":"婚姻狀況","start":25,'length':4})
        attrGroup.append({"name":"現有子女數","start":30,'length':5})
        attrGroup.append({"name":"居住方式","start":36,'length':2})
        attrGroup.append({"name":"縣市","start":39,'length':25})
        
        for group in attrGroup:
            d = {"attrGroup":group["name"]}
            for row in range(group["start"],group["start"]+group["length"]):
                #去掉空白
                attr = "".join(sheet['A'+str(row)].value.split())
                total = sheet['B'+str(row)].value
                d["attr"] = attr
                d["activity"] = "總計"
                d["mainCount"] = total
                d["secondaryCount"] = total
                #print(d)
                util.DataToDB(self.connection,"AgingSurveyActivity",d)
                
                for col in range(4,28,3):
                    #去掉空白
                    activity = "".join(sheet.cell(row=1, column=col).value.split())
                    d["activity"] = activity
                    mainV = sheet.cell(row=row, column=col+1).value
                    if util.IsNumber(mainV):
                        d["mainCount"] = int(total*float(mainV)/100)
                    else:
                        d["mainCount"] = 0
                        
                    secondV = sheet.cell(row=row, column=col+2).value
                    if util.IsNumber(secondV):
                        d["secondaryCount"] = int(total*float(secondV)/100)
                    else:
                        d["secondaryCount"] = 0
                    #print(d)
                    util.DataToDB(self.connection,"AgingSurveyActivity",d)
                    
                d["activity"] = "無參加活動"
                v = sheet.cell(row=row, column=28).value
                if util.IsNumber(v):
                    v = int(total*float(v)/100)
                    d["mainCount"] = v
                    d["secondaryCount"] = v
                else:
                    d["mainCount"] = 0
                    d["secondaryCount"] = 0
                #print(d)
                util.DataToDB(self.connection,"AgingSurveyActivity",d)
    
    def UpdateAgingSurveyFeeling(self):
        print("update AgingSurveyFeeling")
        
        sheet = self.wb["心情感受"]
        attrGroup = []
        attrGroup.append({"name":"總計","start":3,'length':1})
        attrGroup.append({"name":"性別","start":5,'length':2})
        attrGroup.append({"name":"年齡別","start":8,'length':4})
        attrGroup.append({"name":"身分別","start":13,'length':3})
        attrGroup.append({"name":"教育程度","start":17,'length':7})
        attrGroup.append({"name":"婚姻狀況","start":25,'length':4})
        attrGroup.append({"name":"現有子女數","start":30,'length':5})
        attrGroup.append({"name":"居住方式","start":36,'length':2})
        attrGroup.append({"name":"縣市","start":39,'length':25})
        
        for group in attrGroup:
            d = {"attrGroup":group["name"]}
            for row in range(group["start"],group["start"]+group["length"]):
                #去掉空白
                attr = "".join(sheet['A'+str(row)].value.split())
                total = sheet['B'+str(row)].value
                d["attr"] = attr
                d["feel"] = "總計"
                d["neverCount"] = total
                d["sometimesCount"] = total
                d["oftenCount"] = total
                #print(d)
                util.DataToDB(self.connection,"AgingSurveyFeeling",d)
                
                for col in range(4,37,3):
                    #去掉空白
                    feel = "".join(sheet.cell(row=1, column=col).value.split())
                    d["feel"] = feel
                    neverV = sheet.cell(row=row, column=col).value
                    if util.IsNumber(neverV):
                        d["neverCount"] = int(total*float(neverV)/100)
                    else:
                        d["neverCount"] = 0
                        
                    sometimesV = sheet.cell(row=row, column=col+1).value
                    if util.IsNumber(sometimesV):
                        d["sometimesCount"] = int(total*float(sometimesV)/100)
                    else:
                        d["sometimesCount"] = 0
                        
                    oftenV = sheet.cell(row=row, column=col+2).value
                    if util.IsNumber(oftenV):
                        d["oftenCount"] = int(total*float(oftenV)/100)
                    else:
                        d["oftenCount"] = 0
                    #print(d)
                    util.DataToDB(self.connection,"AgingSurveyFeeling",d)
                    
    
    def UpdateAgingSurveyIdealLive(self):
        print("update AgingSurveyIdealLive")
        
        sheet = self.wb["理想居住方式"]
        attrGroup = []
        attrGroup.append({"name":"總計","start":3,'length':1})
        attrGroup.append({"name":"性別","start":5,'length':2})
        attrGroup.append({"name":"年齡別","start":8,'length':4})
        attrGroup.append({"name":"身分別","start":13,'length':3})
        attrGroup.append({"name":"教育程度","start":17,'length':7})
        attrGroup.append({"name":"婚姻狀況","start":25,'length':4})
        attrGroup.append({"name":"現有子女數","start":30,'length':5})
        attrGroup.append({"name":"居住方式","start":36,'length':2})
        attrGroup.append({"name":"縣市","start":39,'length':25})
        
        for group in attrGroup:
            d = {"attrGroup":group["name"]}
            for row in range(group["start"],group["start"]+group["length"]):
                #去掉空白
                attr = "".join(sheet['A'+str(row)].value.split())
                total = sheet['B'+str(row)].value
                d["attr"] = attr
                d["liveType"] = "總計"
                d["count"] = total
                #print(d)
                util.DataToDB(self.connection,"AgingSurveyIdealLive",d)
                
                for col in range(4,11):
                    #去掉空白
                    liveType = "".join(sheet.cell(row=2, column=col).value.split())
                    d["liveType"] = liveType
                    v = sheet.cell(row=row, column=col).value
                    if util.IsNumber(v):
                        d["count"] = int(total*float(v)/100)
                    else:
                        d["count"] = 0
                    #print(d)
                    util.DataToDB(self.connection,"AgingSurveyIdealLive",d)
    
    def UpdateAgingSurveyExpect(self):
        print("update AgingSurveyExpect")
        
        sheet = self.wb["對生活的期望"]
        attrGroup = []
        attrGroup.append({"name":"總計","start":3,'length':1})
        attrGroup.append({"name":"性別","start":5,'length':2})
        attrGroup.append({"name":"年齡別","start":8,'length':4})
        attrGroup.append({"name":"身分別","start":13,'length':3})
        attrGroup.append({"name":"教育程度","start":17,'length':7})
        attrGroup.append({"name":"婚姻狀況","start":25,'length':4})
        attrGroup.append({"name":"現有子女數","start":30,'length':5})
        attrGroup.append({"name":"居住方式","start":36,'length':2})
        attrGroup.append({"name":"縣市","start":39,'length':25})
        
        for group in attrGroup:
            d = {"attrGroup":group["name"]}
            for row in range(group["start"],group["start"]+group["length"]):
                #去掉空白
                attr = "".join(sheet['A'+str(row)].value.split())
                total = sheet['B'+str(row)].value
                d["attr"] = attr
                d["expect"] = "總計"
                d["count"] = total
                #print(d)
                util.DataToDB(self.connection,"AgingSurveyExpect",d)
                
                for col in range(3,17):
                    #去掉空白
                    expect = "".join(sheet.cell(row=2, column=col).value.split())
                    d["expect"] = expect
                    v = sheet.cell(row=row, column=col).value
                    if util.IsNumber(v):
                        d["count"] = int(total*float(v)/100)
                    else:
                        d["count"] = 0
                    #print(d)
                    util.DataToDB(self.connection,"AgingSurveyExpect",d)
    
    def UpdateAgingSurveyLivingDifficulty(self):
        print("update AgingSurveyLivingDifficulty")
        
        sheet = self.wb["起居困難"]
        attrGroup = []
        attrGroup.append({"name":"總計","start":3,'length':1})
        attrGroup.append({"name":"性別","start":5,'length':2})
        attrGroup.append({"name":"年齡別","start":8,'length':4})
        attrGroup.append({"name":"身分別","start":13,'length':3})
        attrGroup.append({"name":"教育程度","start":17,'length':7})
        attrGroup.append({"name":"婚姻狀況","start":25,'length':4})
        attrGroup.append({"name":"現有子女數","start":30,'length':5})
        attrGroup.append({"name":"居住方式","start":36,'length':2})
        attrGroup.append({"name":"縣市","start":39,'length':25})
        
        for group in attrGroup:
            d = {"attrGroup":group["name"]}
            for row in range(group["start"],group["start"]+group["length"]):
                #去掉空白
                attr = "".join(sheet['A'+str(row)].value.split())
                total = sheet['B'+str(row)].value
                d["attr"] = attr
                d["difficulty"] = "總計"
                d["count"] = total
                #print(d)
                util.DataToDB(self.connection,"AgingSurveyLivingDifficulty",d)
                
                for col in range(4,17):
                    #去掉空白
                    difficulty = "".join(sheet.cell(row=2, column=col).value.split())
                    d["difficulty"] = difficulty
                    v = sheet.cell(row=row, column=col).value
                    if util.IsNumber(v):
                        d["count"] = int(total*float(v)/100)
                    else:
                        d["count"] = 0
                    #print(d)
                    util.DataToDB(self.connection,"AgingSurveyLivingDifficulty",d)
    
    def UpdateAgingSurveyFunctionDifficulty(self):
        print("update AgingSurveyFunctionDifficulty")
        
        sheet = self.wb["工具性日常活動"]
        attrGroup = []
        attrGroup.append({"name":"總計","start":3,'length':1})
        attrGroup.append({"name":"性別","start":5,'length':2})
        attrGroup.append({"name":"年齡別","start":8,'length':4})
        attrGroup.append({"name":"身分別","start":13,'length':3})
        attrGroup.append({"name":"教育程度","start":17,'length':7})
        attrGroup.append({"name":"婚姻狀況","start":25,'length':4})
        attrGroup.append({"name":"現有子女數","start":30,'length':5})
        attrGroup.append({"name":"居住方式","start":36,'length':2})
        attrGroup.append({"name":"縣市","start":39,'length':25})
        
        for group in attrGroup:
            d = {"attrGroup":group["name"]}
            for row in range(group["start"],group["start"]+group["length"]):
                #去掉空白
                attr = "".join(sheet['A'+str(row)].value.split())
                total = sheet['B'+str(row)].value
                d["attr"] = attr
                d["difficulty"] = "總計"
                d["noNeedCount"] = total
                d["canCount"] = total
                d["diffcultCount"] = total
                d["canNotCount"] = total
                #print(d)
                util.DataToDB(self.connection,"AgingSurveyFunctionDifficulty",d)
                
                for col in range(4,36,4):
                    #去掉空白
                    difficulty = "".join(sheet.cell(row=1, column=col).value.split())
                    d["difficulty"] = difficulty
                    noNeedV = sheet.cell(row=row, column=col).value
                    if util.IsNumber(noNeedV):
                        d["noNeedCount"] = int(total*float(noNeedV)/100)
                    else:
                        d["noNeedCount"] = 0
                        
                    canV = sheet.cell(row=row, column=col+1).value
                    if util.IsNumber(canV):
                        d["canCount"] = int(total*float(canV)/100)
                    else:
                        d["canCount"] = 0
                        
                    diffcultV = sheet.cell(row=row, column=col+2).value
                    if util.IsNumber(diffcultV):
                        d["diffcultCount"] = int(total*float(diffcultV)/100)
                    else:
                        d["diffcultCount"] = 0
                        
                    canNotV = sheet.cell(row=row, column=col+3).value
                    if util.IsNumber(canNotV):
                        d["canNotCount"] = int(total*float(canNotV)/100)
                    else:
                        d["canNotCount"] = 0
                    #print(d)
                    util.DataToDB(self.connection,"AgingSurveyFunctionDifficulty",d)

    def UpdateData(self):
        self.UpdateAgingSurveyAge()
        self.UpdateAgingSurveyHouseType()
        self.UpdateAgingSurveyFamilyType()
        self.UpdateAgingSurveyChronic()
        self.UpdateAgingSurveyActivity()
        self.UpdateAgingSurveyFeeling()
        self.UpdateAgingSurveyIdealLive()
        self.UpdateAgingSurveyExpect()
        self.UpdateAgingSurveyLivingDifficulty()
        self.UpdateAgingSurveyFunctionDifficulty()
                
      