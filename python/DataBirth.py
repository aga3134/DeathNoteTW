# -*- coding: utf-8 -*-
"""
Created on Wed Dec 27 17:51:03 2017

@author: aga
"""

from openpyxl import load_workbook
import Util as util
import re

class DataBirth:
    def __init__(self, connection):
        self.connection = connection
        
    def CreateTable(self):
        print("Create Table for DataBirth")
        with self.connection.cursor() as cursor:
            sql = "CREATE TABLE IF NOT EXISTS PopulationByAge (\
                year INT(11),\
                county VARCHAR(255),\
                sex VARCHAR(10),\
                minAge INT(11),\
                maxAge INT(11),\
                count INT(11),\
                PRIMARY KEY (year,county,sex,minAge,maxAge)\
                );"
            cursor.execute(sql)
            
            sql = "CREATE TABLE IF NOT EXISTS MarriageByAge (\
                year INT(11),\
                status VARCHAR(255),\
                sex VARCHAR(10),\
                minAge INT(11),\
                maxAge INT(11),\
                count INT(11),\
                PRIMARY KEY (year,status,sex,minAge,maxAge)\
                );"
            cursor.execute(sql)
            
            sql = "CREATE TABLE IF NOT EXISTS BirthByCounty (\
                year INT(11),\
                county VARCHAR(255),\
                maleBirth INT(11),\
                femaleBirth INT(11),\
                maleDeath INT(11),\
                femaleDeath INT(11),\
                socialIncrease INT(11),\
                marriage  INT(11),\
                divorce INT(11),\
                PRIMARY KEY (year,county)\
                );"
            cursor.execute(sql)
            
            sql = "CREATE TABLE IF NOT EXISTS BirthCount (\
                year INT(11),\
                county VARCHAR(255),\
                minMonAge INT(11),\
                maxMonAge INT(11),\
                count INT(11),\
                PRIMARY KEY (year,county,minMonAge,maxMonAge)\
                );"
            cursor.execute(sql)
            
            sql = "CREATE TABLE IF NOT EXISTS FertilityRateByAge (\
                year INT(11),\
                county VARCHAR(255),\
                minAge INT(11),\
                maxAge INT(11),\
                rate FLOAT,\
                PRIMARY KEY (year,county,minAge,maxAge)\
                );"
            cursor.execute(sql)
            
            sql = "CREATE TABLE IF NOT EXISTS ExpectLife (\
                year INT(11),\
                sex VARCHAR(10),\
                age INT(11),\
                life FLOAT,\
                PRIMARY KEY (year,sex,age)\
                );"
            cursor.execute(sql)
            
            sql = "CREATE TABLE IF NOT EXISTS PopulationProjection (\
                year INT(11),\
                sex VARCHAR(10),\
                age INT(11),\
                estimateParam VARCHAR(255),\
                count INT(11),\
                PRIMARY KEY (year,sex,age,estimateParam)\
                );"
            cursor.execute(sql)
            
            sql = "CREATE TABLE IF NOT EXISTS ProjectionIndex (\
                year INT(11),\
                estimateParam VARCHAR(255),\
                maleCount INT(11),\
                femaleCount INT(11),\
                maleLife FLOAT,\
                femaleLife FLOAT,\
                totalFertilityRate FLOAT,\
                birthCount INT(11),\
                deathCount INT(11),\
                socialIncrease INT(11),\
                PRIMARY KEY (year,estimateParam)\
                );"
            cursor.execute(sql)

        self.connection.commit()
        
    def UpdatePopulationByAge(self):
        wb = load_workbook(filename = 'data/生之章/各縣市人口數按性別及五歲年齡組分(63~105).xlsx')
        for year in range(63,106,1):
            print(year)
            sheet = wb[str(year)]
            sexArr = ["男","女"]
            for row in range(6,93,3):
                countyCell = sheet['A'+str(row)]
                if(countyCell.value is None):
                    continue
                #去掉空白
                county = "".join(countyCell.value.split())
                if(county == ""):
                    continue
                print(county)
                
                for sexIndex in range(len(sexArr)):
                    sex = sexArr[sexIndex]
                    sexRow = row+sexIndex
                    #print(sex)
    
                    #1~4歲分別計
                    countArr = []
                    countArr.append(sheet['D'+str(sexRow)].value)
                    countArr.append(sheet['F'+str(sexRow)].value)
                    countArr.append(sheet['G'+str(sexRow)].value)
                    countArr.append(sheet['H'+str(sexRow)].value)
                    countArr.append(sheet['I'+str(sexRow)].value)
                    
                    d = {"year":year+1911, "county":county, "sex": sex}
                    for i in range(len(countArr)):
                        if(countArr[i] is None):
                            continue
                        d["minAge"] = d["maxAge"] = i
                        d["count"] = countArr[i]
                        #print(d)
                        util.DataToDB(self.connection,"PopulationByAge",d) 
                    #總計
                    d["minAge"] = 0
                    d["maxAge"] = 100
                    d["count"] = sheet['C'+str(sexRow)].value
                    util.DataToDB(self.connection,"PopulationByAge",d)
                    #0~4歲總計
                    if(countArr[0] is not None):
                        d["minAge"] = 0
                        d["maxAge"] = 4
                        d["count"] = countArr[0]+sheet['E'+str(sexRow)].value
                        util.DataToDB(self.connection,"PopulationByAge",d)
                    #5歲年齡組
                    for col in range(1,20,1):
                        age = re.findall(r"\d+",sheet.cell(row=3, column=9+col).value)
                        v = sheet.cell(row=sexRow, column=9+col).value
                        if(v is None):
                            continue
                        d["minAge"] = int(age[0])
                        d["maxAge"] = int(age[1])
                        d["count"] = v
                        #print(d)
                        util.DataToDB(self.connection,"PopulationByAge",d)
                    
                    v = sheet['AC'+str(sexRow)].value
                    if(v is None):
                        continue
                    d["minAge"] = d["maxAge"] = 100
                    d["count"] = v
                    #print(d)
                    util.DataToDB(self.connection,"PopulationByAge",d)
                    
    def UpdateMarriageByAge(self):
        wb = load_workbook(filename = 'data/生之章/現住人口數按性別、年齡及婚姻狀況分(96~105).xlsx')
        for year in range(96,106,1):
            print(year)
            sheet = wb[str(year)+".1"]
            sexArr = {}
            sexArr["男"] = 1
            sexArr["女"] = 2
            statusArr = {}
            statusArr["總計"] = 2
            statusArr["未婚"] = 5
            statusArr["有偶"] = 8
            statusArr["離婚"] = 11
            statusArr["喪偶"] = 14
            
            for status in statusArr:
                for sex in sexArr:
                    d = {"year":year+1911, "status":status, "sex": sex}
                    d["minAge"] = 0
                    d["maxAge"] = 14
                    col = statusArr[status]+sexArr[sex]
                    v = sheet.cell(row=6, column=col).value
                    d["count"] = v
                    #print(d)
                    util.DataToDB(self.connection,"MarriageByAge",d)
            
                    for row in range(7,24,1):
                        age = re.findall(r"\d+",sheet['A'+str(row)].value)
                        d["minAge"] = int(age[0])
                        d["maxAge"] = int(age[1])
                        v = sheet.cell(row=row, column=col).value
                        d["count"] = v
                        #print(d)
                        util.DataToDB(self.connection,"MarriageByAge",d)
                        
                    d["minAge"] = 100
                    d["maxAge"] = 100
                    v = sheet.cell(row=24, column=col).value
                    d["count"] = v
                    #print(d)
                    util.DataToDB(self.connection,"MarriageByAge",d)
                        
    
    def UpdateBirthByCounty(self):
        wb = load_workbook(filename = 'data/生之章/出生死亡統計.xlsx')
        for year in range(1992,2017,1):
            print(year)
            sheet = wb[str(year)]
            
            for row in range(8,36,1):
                countyCell = sheet['A'+str(row)]
                if(countyCell.value is None):
                    continue
                #去掉英文留中文
                words = countyCell.value.split()
                arr = []
                for word in words:
                    if(not re.findall(r"[a-zA-Z]",word)):
                        arr.append(word)
                county = "".join(arr)
                #只留有縣跟市的row
                if(not re.findall(r"[縣市]",county)):
                    continue
                #print(county)
                
                maleBirth = sheet['D'+str(row)].value
                femaleBirth = sheet['E'+str(row)].value
                maleDeath = sheet['H'+str(row)].value
                femaleDeath = sheet['I'+str(row)].value
                socialIncrease = sheet['J'+str(row)].value
                marriage = sheet['P'+str(row)].value
                divorce = sheet['R'+str(row)].value
                
                d = {"year":year, "county":county, 
                     "maleBirth": maleBirth,"femaleBirth":femaleBirth,
                     "maleDeath":maleDeath,"femaleDeath":femaleDeath,
                     "socialIncrease":socialIncrease,"marriage":marriage,
                     "divorce":divorce}
                #print(d)
                util.DataToDB(self.connection,"BirthByCounty",d)
            
    
    def UpdateBirthCount(self):
        wb = load_workbook(filename = 'data/生之章/各縣市嬰兒出生數按生母年齡分(96~105).xlsx')
        for year in range(96,106,1):
            print(year)
            sheet = wb[str(year)]
            
            for row in range(6,36,1):
                countyCell = sheet['A'+str(row)]
                if(countyCell.value is None):
                    continue

                if(countyCell.value is None):
                    continue
                #去掉空白
                county = "".join(countyCell.value.split())
                if(county == ""):
                    continue
                #只留有縣跟市的row
                if(not re.findall(r"[縣市]",county)):
                    continue
                #print(county)
                
                d = {"year":year+1911, "county":county}
                d["minMonAge"] = 0
                d["maxMonAge"] = 20
                v = sheet.cell(row=row, column=5).value
                d["count"] = v
                #print(d)
                util.DataToDB(self.connection,"BirthCount",d)
        
                for col in range(6,11,1):
                    ageRow = 5
                    if(year < 103):
                        ageRow = 7
                    age = re.findall(r"\d+",sheet.cell(row=ageRow, column=col).value)
                    d["minMonAge"] = int(age[0])
                    d["maxMonAge"] = int(age[1])
                    v = sheet.cell(row=row, column=col).value
                    d["count"] = v
                    #print(d)
                    util.DataToDB(self.connection,"BirthCount",d)
                    
                d["minMonAge"] = 45
                d["maxMonAge"] = 100
                v = sheet.cell(row=row, column=11).value
                d["count"] = v
                #print(d)
                util.DataToDB(self.connection,"BirthCount",d)
    
    def UpdateFertilityRateByAge(self):
        wb = load_workbook(filename = 'data/生之章/生育率.xlsx')
        for year in range(1999,2017,1):
            print(year)
            sheet = wb[str(year)]
            
            for row in range(8,36,1):
                countyCell = sheet['A'+str(row)]
                if(countyCell.value is None):
                    continue
                #去掉空白
                county = "".join(countyCell.value.split())
                if(county == ""):
                    continue
                #只留有縣跟市的row
                if(not re.findall(r"[縣市]",county)):
                    continue
                #print(county)
                
                d = {"year":year, "county":county}
                for col in range(4,11,1):
                    age = re.findall(r"\d+",sheet.cell(row=5, column=col).value)
                    d["minAge"] = int(age[0])
                    d["maxAge"] = int(age[1])
                    v = sheet.cell(row=row, column=col).value
                    d["rate"] = v
                    #print(d)
                    util.DataToDB(self.connection,"FertilityRateByAge",d)
    
    def UpdateExpectLife(self):
        wb = load_workbook(filename = 'data/生之章/平均餘命.xlsx')
        sheet = wb["Sheet1"]
        for row in range(7,72,1):
            #取後四碼年份
            year = sheet['A'+str(row)].value[-4:]
            print(year)
            
            d = {"year":year}
            sexArr = {}
            sexArr["全部"] = 2
            sexArr["男"] = 11
            sexArr["女"] = 20
            for sex in sexArr:
                d["sex"] = sex
                for col in range(0,9,1):
                    c = sexArr[sex]+col
                    age = re.findall(r"\d+",sheet.cell(row=5, column=c).value)
                    d["age"] = int(age[0])
                    v = sheet.cell(row=row, column=c).value
                    d["life"] = v
                    #print(d)
                    util.DataToDB(self.connection,"ExpectLife",d)
    
    def UpdatePopulationProjection(self):
        def UpdateProjectionData(wb,estimateParam):
            sexArr = ["男","女"]
            for sex in sexArr:
                sheet = wb["單一年齡人口-"+sex]
                for row in range(3,49,1):
                    year = sheet['A'+str(row)].value
                    print(year)
                    
                    for col in range(4,105,1):
                        d = {"year":year, "sex":sex, "age":col-4,"estimateParam":estimateParam}
                        v = sheet.cell(row=row, column=col).value
                        d["count"] = v
                        #print(d)
                        util.DataToDB(self.connection,"PopulationProjection",d)
            
        wb = load_workbook(filename = 'data/生之章/中華民國人口推計（105至150年）數據－低推估.xlsx')
        UpdateProjectionData(wb,"低推估")
        
        wb = load_workbook(filename = 'data/生之章/中華民國人口推計（105至150年）數據－中推計.xlsx')
        UpdateProjectionData(wb,"中推估")
        
        wb = load_workbook(filename = 'data/生之章/中華民國人口推計（105至150年）數據－高推估.xlsx')
        UpdateProjectionData(wb,"高推估")
    
    def UpdateProjectionIndex(self):
        def UpdateIndexData(wb,estimateParam):
            sheetIndex = wb["重要指標"]
            sheetPop = wb["人口變動"]
            for row in range(5,51,1):
                year = sheetIndex['A'+str(row)].value
                print(year)
                
                maleCount = sheetIndex['D'+str(row)].value
                femaleCount = sheetIndex['E'+str(row)].value
                totalFertilityRate  = sheetIndex['G'+str(row)].value
                maleLife  = sheetIndex['H'+str(row)].value
                femaleLife  = sheetIndex['I'+str(row)].value
                birthCount = sheetPop['G'+str(row)].value
                deathCount = sheetPop['I'+str(row)].value
                socialIncrease = sheetPop['K'+str(row)].value
                d = {}
                d["year"] = year
                d["estimateParam"] = estimateParam
                d["maleCount"] = maleCount
                d["femaleCount"] = femaleCount
                d["totalFertilityRate"] = totalFertilityRate
                d["maleLife"] = maleLife
                d["femaleLife"] = femaleLife
                d["birthCount"] = birthCount
                d["deathCount"] = deathCount
                d["socialIncrease"] = socialIncrease
                
                #print(d)
                util.DataToDB(self.connection,"ProjectionIndex",d)
            
        wb = load_workbook(filename = 'data/生之章/中華民國人口推計（105至150年）數據－低推估.xlsx')
        UpdateIndexData(wb,"低推估")
        
        wb = load_workbook(filename = 'data/生之章/中華民國人口推計（105至150年）數據－中推計.xlsx')
        UpdateIndexData(wb,"中推估")
        
        wb = load_workbook(filename = 'data/生之章/中華民國人口推計（105至150年）數據－高推估.xlsx')
        UpdateIndexData(wb,"高推估")
            
    def UpdateData(self):
        self.UpdatePopulationByAge()
        self.UpdateMarriageByAge()
        self.UpdateBirthByCounty()
        self.UpdateBirthCount()
        self.UpdateFertilityRateByAge()
        self.UpdateExpectLife()
        self.UpdatePopulationProjection()
        self.UpdateProjectionIndex()
                
      